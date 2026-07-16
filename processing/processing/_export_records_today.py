import pandas as pd
import psycopg2
import datetime
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph

for folder in ['export/records']:
    os.makedirs(folder, exist_ok=True)

start_time = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
global_filename = f'export/records/records_export_{start_time}'

DB_CONFIG = {
    'host': 'localhost',
    'port': '5432',
    'dbname': 'hotel_breakfast',
    'user': 'postgres',
    'password': ''
}


def get_db_connection():
    return psycopg2.connect(**DB_CONFIG)


def select_records():
    conn = get_db_connection()
    cursor = conn.cursor()

    query = """
        with count_coms as (
            select c.guest_id, count(c.comment_id) as count from comments c
            group by 1
        )
        select
            row_number() over (order by r.room_number)  as record,
            r.room_number as room,
            g.profile_id,
            r.reservation_id as res_id,
            g.f_name as name,
            case
                when g.birth_date is null then 'pass'
                when extract(month from g.birth_date) = extract(month from localtimestamp)
                 and extract(day from g.birth_date) = extract(day from localtimestamp) then 'yes'
                else ''
            end as bday,
            -- r.title,
            case
                when r.title in ('mr')
                    then 'male'
                when r.title in ('mrs', 'ms')
                    then 'female'
            end as gender,
            UPPER(r.nationality_code) || ' (' || COALESCE(r.nationality_code_description, '') || ') (' || COALESCE(r.language, '') || ')' as nationality,
            (
                select
                    coalesce(sum(r2.adult_count + r2.child_count), 0) || ' (' ||
                    coalesce(sum(r2.child_count), 0) || ')'
                from records r2
                where date_trunc('day', r2.created_at) = date_trunc('day', localtimestamp)
                and r2.room_number = r.room_number
                and case
                        when r.reservation_status in ('checked in', 'due out', 'walk in')
                            then r2.reservation_status in ('checked in', 'due out', 'walk in')
                        when r.reservation_status in ('due in', 'no show')
                            then r2.reservation_status in ('due in', 'no show')
                    end
            ) as guest_count,
            r.reservation_status as res_status,
            r.arrival_date || '  ' || coalesce(r.arrival_time, '--:--') as arrival,
            r.departure_date || '  ' || coalesce(r.departure_time, '--:--') as departure,
            r.vip_code || ' (' || r.vip_code_description || ')' as vip_status,
            -- r.membership_level_tng as tng,
            r.room_type || ' (' || COALESCE(r.room_class, '') || ')' as type,
            cc.count::int as coms,
            r.attended_at as check_time
        from records r
        inner join guests g on r.guest_id = g.guest_id
        left join count_coms cc on r.guest_id = cc.guest_id
        where date_trunc('day', r.created_at) = date_trunc('day', localtimestamp)
        order by r.record_id;
    """

    try:
        cursor.execute(query)
        rows = cursor.fetchall()
        col_names = [desc[0] for desc in cursor.description]

        df = pd.DataFrame(rows, columns=col_names)
        print(f'Выбрано {len(df)} записей из records')
        return df
    except Exception as e:
        print(f'Ошибка при выполнении SELECT: {e}')
        raise
    finally:
        cursor.close()
        conn.close()


def export_to_csv(df, filename=None):
    if filename is None:
        filename = global_filename + '.csv'

    try:
        df.to_csv(filename, sep=';', index=False, encoding='utf-8-sig')
        print(f'Данные сохранены в CSV: {filename}')
    except Exception as e:
        print(f'Ошибка при сохранении CSV: {e}')
        raise


def export_to_excel(df, filename=None):
    if filename is None:
        filename = global_filename + '.xlsx'

    try:
        df.to_excel(filename, index=False, engine='openpyxl')

        wb = load_workbook(filename)
        ws = wb.active
        col_names = df.columns.tolist()

        for idx, col_name in enumerate(col_names, start=1):
            col_letter = get_column_letter(idx)

            width = fmt.get('width')
            align = fmt.get('align', 'left') # 'left', 'center', 'right'
            valign = fmt.get('valign', 'center') # 'top', 'center', 'bottom'

            if width is None:
                max_length = 0
                for row in ws.iter_rows(min_col=idx, max_col=idx):
                    for cell in row:
                        if cell.value is not None:
                            length = len(str(cell.value))
                            if length > max_length:
                                max_length = length
                width = max_length + 2

            ws.column_dimensions[col_letter].width = width

            for row in ws.iter_rows(min_col=idx, max_col=idx):
                for cell in row:
                    if align == 'left':
                        cell.alignment = Alignment(horizontal='left', vertical=valign)
                    elif align == 'center':
                        cell.alignment = Alignment(horizontal='center', vertical=valign)
                    elif align == 'right':
                        cell.alignment = Alignment(horizontal='right', vertical=valign)
                    else:
                        cell.alignment = Alignment(horizontal='general', vertical=valign)

        wb.save(filename)
        print(f'Данные сохранены в Excel: {filename}')
    except Exception as e:
        print(f'Ошибка при сохранении Excel: {e}')
        raise


def export_to_pdf(df, filename=None, page_size=A4, orientation='portrait', column_widths=None):
    if filename is None:
        filename = global_filename + '.pdf'

    if orientation == 'landscape':
        page_size = landscape(page_size)

    doc = SimpleDocTemplate(filename, pagesize=page_size,
                            leftMargin=0.3 * inch, rightMargin=0.3 * inch,
                            topMargin=0.3 * inch, bottomMargin=0.3 * inch)

    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Heading4'],
        alignment=TA_CENTER,
        fontSize=10,
        textColor=colors.white,
        backColor=colors.grey,
        leading=12
    )
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        alignment=TA_CENTER,
        fontSize=9,
        leading=11
    )

    headers = df.columns.tolist()
    header_cells = [Paragraph(str(h), header_style) for h in headers]

    data_rows = []

    max_lengths = [len(str(h)) for h in headers]
    for _, row in df.iterrows():
        row_cells = []
        for i, val in enumerate(row):
            if pd.isna(val):
                text = ''
            elif isinstance(val, (pd.Timestamp, datetime.datetime)):
                text = val.strftime('%Y-%m-%d %H:%M')
            else:
                text = str(val)
            row_cells.append(Paragraph(text, cell_style))

            length = len(text)
            if length > max_lengths[i]:
                max_lengths[i] = length
        data_rows.append(row_cells)

    table_data = [header_cells] + data_rows

    char_width_pt = 6.5
    if column_widths is None:
        col_widths = [max_len * char_width_pt + 10 for max_len in max_lengths]
    else:
        if not isinstance(column_widths, dict):
            raise ValueError("column_widths должен быть словарём {имя_колонки: ширина_в_пунктах}")
        col_widths = []
        for idx, col_name in enumerate(headers):
            if col_name in column_widths:
                col_widths.append(column_widths[col_name])
            else:
                col_widths.append(max_lengths[idx] * char_width_pt + 10)

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
    ])

    # Добавляем чередование цветов для строк (начиная с 1, т.к. 0 - это заголовок)
    for i in range(1, len(data_rows) + 1):
        if i % 2 == 1:  # Нечетные строки - светло-серые
            style.add('BACKGROUND', (0, i), (-1, i), colors.lightgrey)
            style.add('GRID', (0, i), (-1, i), 0.5, colors.black)
        # Четные строки остаются белыми (можно задать явно)
        else:
            style.add('BACKGROUND', (0, i), (-1, i), colors.white)
            style.add('GRID', (0, i), (-1, i), 0.5, colors.black)

    table.setStyle(style)
    doc.build([table])
    print(f'Данные сохранены в PDF: {filename}')
    return filename


def main():
    df = select_records()

    if df.empty:
        return

    int_columns = ['room_n', 'profile_id', 'res_id', 'coms']  # и другие, где нужны целые числа
    for col in int_columns:
        if col in df.columns:
            df[col] = df[col].fillna(0).astype(int)

    # csv
    # export_to_csv(df)

    # xslx
    # export_to_excel(df)

    # pdf
    pdf_column_widths = {
        'record': 40,
        'room': 35,
        'profile_id': 55,
        'res_id': 50,
        'name': 80,
        'bday': 35,
        'gender': 45,
        'nationality': 70,
        'guest_count': 45,
        'res_status': 60,
        'arrival': 60,
        'departure': 60,
        'vip_status': 60,
        'type': 40,
        'coms': 35,
        'check_time': 65,
    }

    export_to_pdf(df, orientation='landscape', column_widths=pdf_column_widths)

if __name__ == '__main__':
    main()