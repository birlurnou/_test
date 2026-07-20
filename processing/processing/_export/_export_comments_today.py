import pandas as pd
import psycopg2
import datetime
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.fonts import addMapping

for folder in ['export/comments']:
    os.makedirs(folder, exist_ok=True)

start_time = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
global_filename = f'export/comments/comments_export_{start_time}'

DB_CONFIG = {
    'host': 'localhost',
    'port': '5432',
    'dbname': 'hotel_breakfast',
    'user': 'postgres',
    'password': ''
}


def get_db_connection():
    return psycopg2.connect(**DB_CONFIG)


def select_comments():
    conn = get_db_connection()
    cursor = conn.cursor()

    query = """
        select
            g.f_name as name,
            r.room_number as room,
            c.comment,
            c.created_at
        from comments c
        inner join guests g on c.guest_id = g.guest_id
        inner join records r on c.guest_id = r.guest_id
        where date_trunc('day', r.created_at) = date_trunc('day', localtimestamp)
        order by room asc, created_at desc;
    """

    try:
        cursor.execute(query)
        rows = cursor.fetchall()
        col_names = [desc[0] for desc in cursor.description]

        df = pd.DataFrame(rows, columns=col_names)
        print(f'Выбрано {len(df)} записей из records')
        return df
    except Exception as e:
        print(f'Ошибка при выполнении SELECT: {e}')
        raise
    finally:
        cursor.close()
        conn.close()


def export_to_csv(df, filename=None):
    if filename is None:
        filename = global_filename + '.csv'

    try:
        df.to_csv(filename, sep=';', index=False, encoding='utf-8-sig')
        print(f'Данные сохранены в CSV: {filename}')
    except Exception as e:
        print(f'Ошибка при сохранении CSV: {e}')
        raise


def export_to_excel(df, filename=None):
    if filename is None:
        filename = global_filename + '.xlsx'

    try:
        df.to_excel(filename, index=False, engine='openpyxl')

        wb = load_workbook(filename)
        ws = wb.active
        col_names = df.columns.tolist()

        for idx, col_name in enumerate(col_names, start=1):
            col_letter = get_column_letter(idx)

            width = fmt.get('width')
            align = fmt.get('align', 'left') # 'left', 'center', 'right'
            valign = fmt.get('valign', 'center') # 'top', 'center', 'bottom'

            if width is None:
                max_length = 0
                for row in ws.iter_rows(min_col=idx, max_col=idx):
                    for cell in row:
                        if cell.value is not None:
                            length = len(str(cell.value))
                            if length > max_length:
                                max_length = length
                width = max_length + 2

            ws.column_dimensions[col_letter].width = width

            for row in ws.iter_rows(min_col=idx, max_col=idx):
                for cell in row:
                    if align == 'left':
                        cell.alignment = Alignment(horizontal='left', vertical=valign)
                    elif align == 'center':
                        cell.alignment = Alignment(horizontal='center', vertical=valign)
                    elif align == 'right':
                        cell.alignment = Alignment(horizontal='right', vertical=valign)
                    else:
                        cell.alignment = Alignment(horizontal='general', vertical=valign)

        wb.save(filename)
        print(f'Данные сохранены в Excel: {filename}')
    except Exception as e:
        print(f'Ошибка при сохранении Excel: {e}')
        raise


def export_to_pdf(df, filename=None, page_size=A4, orientation='portrait', column_widths=None):
    if filename is None:
        filename = global_filename + '.pdf'

    if orientation == 'landscape':
        page_size = landscape(page_size)

    doc = SimpleDocTemplate(filename, pagesize=page_size,
                            leftMargin=0.3 * inch, rightMargin=0.3 * inch,
                            topMargin=0.3 * inch, bottomMargin=0.3 * inch)

    pdfmetrics.registerFont(TTFont('DejaVuSans', 'DejaVuSans.ttf'))
    pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', 'DejaVuSans-Bold.ttf'))

    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Heading4'],
        alignment=TA_CENTER,
        fontSize=10,
        textColor=colors.white,
        backColor=colors.grey,
        leading=12,
        fontName='DejaVuSans'
    )
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        alignment=TA_CENTER,
        fontSize=9,
        leading=11,
        fontName='DejaVuSans'
    )

    headers = df.columns.tolist()
    header_cells = [Paragraph(str(h), header_style) for h in headers]

    data_rows = []

    max_lengths = [len(str(h)) for h in headers]
    for _, row in df.iterrows():
        row_cells = []
        for i, val in enumerate(row):
            if pd.isna(val):
                text = ''
            elif isinstance(val, (pd.Timestamp, datetime.datetime)):
                text = val.strftime('%Y-%m-%d %H:%M')
            else:
                text = str(val)
            row_cells.append(Paragraph(text, cell_style))

            length = len(text)
            if length > max_lengths[i]:
                max_lengths[i] = length
        data_rows.append(row_cells)

    table_data = [header_cells] + data_rows

    char_width_pt = 6.5
    if column_widths is None:
        col_widths = [max_len * char_width_pt + 10 for max_len in max_lengths]
    else:
        if not isinstance(column_widths, dict):
            raise ValueError("column_widths должен быть словарём {имя_колонки: ширина_в_пунктах}")
        col_widths = []
        for idx, col_name in enumerate(headers):
            if col_name in column_widths:
                col_widths.append(column_widths[col_name])
            else:
                col_widths.append(max_lengths[idx] * char_width_pt + 10)

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
    ])

    table.setStyle(style)
    doc.build([table])
    print(f'Данные сохранены в PDF: {filename}')
    return filename


def main():
    df = select_comments()

    if df.empty:
        return

    int_columns = ['room_n', 'profile_id', 'res_id', 'coms']
    for col in int_columns:
        if col in df.columns:
            df[col] = df[col].fillna(0).astype(int)

    # export_to_csv(df)
    # export_to_excel(df)

    pdf_column_widths = {
        'name': 150,
        'room': 50,
        'comment': 400,
        'created_at': 100,
    }
    export_to_pdf(df, orientation='landscape', column_widths=pdf_column_widths)

if __name__ == '__main__':
    main()